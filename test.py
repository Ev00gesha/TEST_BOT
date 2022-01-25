import telebot
import datetime
import config
from telebot import types
from openpyxl import Workbook, load_workbook

bot = telebot.TeleBot(config.TOKEN)

@bot.message_handler(commands=['test'])
def test_time(message):
    time = str(datetime.datetime.today())[11:16]
    spis_time_day = time.split(":")
    for i in range(len(spis_time_day)):
        spis_time_day[i] = int(spis_time_day[i])
    spis_time_day[0] += 3
    
    wb = load_workbook("testchd.xlsx")
    sheet = wb["TEST"]

    for i in range(sheet.max_row):
        spis_time_shd = sheet["C" + str(i + 1)].value.split(":")
        if int(spis_time_day[0]) <= int(spis_time_shd[0] and int(spis_time_day[1]) < int(spis_time_shd[1])):
            bot.send_message(
                message.chat.id,
                f"Следующая игра команды TEST\nДата: {sheet['A' + str(i + 1)].value}\nВремя: {sheet['C' + str(i + 1)].value}\nПротив команды {sheet['B' + str(i + 1)].value}"
            )
            break
        else:
            continue
        
bot.polling(none_stop=True, interval=0)